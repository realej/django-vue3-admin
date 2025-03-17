# -*- coding: utf-8 -*-
import datetime
from urllib.parse import quote

from django.db import transaction
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter, quote_sheetname
from openpyxl.worksheet.table import Table, TableStyleInfo
from rest_framework.decorators import action
from rest_framework.request import Request

from dvadmin.utils.import_export import import_to_data
from dvadmin.utils.json_response import DetailResponse, SuccessResponse
from dvadmin.utils.request_util import get_verbose_name
from dvadmin.system.tasks import async_export_data
from dvadmin.system.models import DownloadCenter


class ImportSerializerMixin:
    """
    Custom import templates、Import function
    """

    # Import fields
    import_field_dict = {}
    # Import the serializer
    import_serializer_class = None
    # Maximum width of table header，default50Characters
    export_column_width = 50

    def is_number(self,num):
        try:
            float(num)
            return True
        except ValueError:
            pass

        try:
            import unicodedata
            unicodedata.numeric(num)
            return True
        except (TypeError, ValueError):
            pass
        return False

    def get_string_len(self, string):
        """
        Get the maximum length of the string
        :param string:
        :return:
        """
        length = 4
        if string is None:
            return length
        if self.is_number(string):
            return length
        for char in string:
            length += 2.1 if ord(char) > 256 else 1
        return round(length, 1) if length <= self.export_column_width else self.export_column_width

    @action(methods=['get','post'],detail=False)
    @transaction.atomic  # Django Transactions,Prevent errors
    def import_data(self, request: Request, *args, **kwargs):
        """
        Import templates
        :param request:
        :param args:
        :param kwargs:
        :return:
        """
        assert self.import_field_dict, "'%s' Please configure the corresponding export template field。" % self.__class__.__name__
        # Export templates
        if request.method == "GET":
            # Sample data
            queryset = self.filter_queryset(self.get_queryset())
            # Exportexcel surface
            response = HttpResponse(content_type="application/msexcel")
            response["Access-Control-Expose-Headers"] = f"Content-Disposition"
            response[
                "Content-Disposition"
            ] = f'attachment;filename={quote(str(f"Import{get_verbose_name(queryset)}template.xlsx"))}'
            wb = Workbook()
            ws1 = wb.create_sheet("data", 1)
            ws1.sheet_state = "hidden"
            ws = wb.active
            row = get_column_letter(len(self.import_field_dict) + 1)
            column = 10
            header_data = [
                "Serial number",
            ]
            validation_data_dict = {}
            for index, ele in enumerate(self.import_field_dict.values()):
                if isinstance(ele, dict):
                    header_data.append(ele.get("title"))
                    choices = ele.get("choices", {})
                    if choices.get("data"):
                        data_list = []
                        data_list.extend(choices.get("data").keys())
                        validation_data_dict[ele.get("title")] = data_list
                    elif choices.get("queryset") and choices.get("values_name"):
                        data_list = choices.get("queryset").values_list(choices.get("values_name"), flat=True)
                        validation_data_dict[ele.get("title")] = list(data_list)
                    else:
                        continue
                    column_letter = get_column_letter(len(validation_data_dict))
                    dv = DataValidation(
                        type="list",
                        formula1=f"{quote_sheetname('data')}!${column_letter}$2:${column_letter}${len(validation_data_dict[ele.get('title')]) + 1}",
                        allow_blank=True,
                    )
                    ws.add_data_validation(dv)
                    dv.add(f"{get_column_letter(index + 2)}2:{get_column_letter(index + 2)}1048576")
                else:
                    header_data.append(ele)
            # Add a column of data
            ws1.append(list(validation_data_dict.keys()))
            for index, validation_data in enumerate(validation_data_dict.values()):
                for inx, ele in enumerate(validation_data):
                    ws1[f"{get_column_letter(index + 1)}{inx + 2}"] = ele
            # Insert export template official data
            df_len_max = [self.get_string_len(ele) for ele in header_data]
            ws.append(header_data)
            # 　Update column width
            for index, width in enumerate(df_len_max):
                ws.column_dimensions[get_column_letter(index + 1)].width = width
            tab = Table(displayName="Table1", ref=f"A1:{row}{column}")  # Name Manager
            style = TableStyleInfo(
                name="TableStyleLight11",
                showFirstColumn=True,
                showLastColumn=True,
                showRowStripes=True,
                showColumnStripes=True,
            )
            tab.tableStyleInfo = style
            ws.add_table(tab)
            wb.save(response)
            return response
        else:
            # fromexcelThe corresponding data structure of the organization，Then save using the serializer
            queryset = self.filter_queryset(self.get_queryset())
            # Get many to many fields
            m2m_fields = [
                ele.name
                for ele in queryset.model._meta.get_fields()
                if hasattr(ele, "many_to_many") and ele.many_to_many == True
            ]
            import_field_dict = {'id':'Update the primary key(Do not change)',**self.import_field_dict}
            data = import_to_data(request.data.get("url"), import_field_dict, m2m_fields)
            for ele in data:
                filter_dic = {'id':ele.get('id')}
                instance = filter_dic and queryset.filter(**filter_dic).first()
                # print(156,ele)
                serializer = self.import_serializer_class(instance, data=ele, request=request)
                serializer.is_valid(raise_exception=True)
                serializer.save()
            return DetailResponse(msg=f"Import successfully！")

    @action(methods=['get'],detail=False)
    def update_template(self,request):
        queryset = self.filter_queryset(self.get_queryset())
        assert self.import_field_dict, "'%s' Please configure the corresponding import template field。" % self.__class__.__name__
        assert self.import_serializer_class, "'%s' Please configure the corresponding import serializer。" % self.__class__.__name__
        data = self.import_serializer_class(queryset, many=True, request=request).data
        # Exportexcel surface
        response = HttpResponse(content_type="application/msexcel")
        response["Access-Control-Expose-Headers"] = f"Content-Disposition"
        response["content-disposition"] = f'attachment;filename={quote(str(f"Export{get_verbose_name(queryset)}.xlsx"))}'
        wb = Workbook()
        ws1 = wb.create_sheet("data", 1)
        ws1.sheet_state = "hidden"
        ws = wb.active
        import_field_dict = {}
        header_data = ["Serial number","Update the primary key(Do not change)"]
        hidden_header = ["#","id"]
        #----Setting options----
        validation_data_dict = {}
        for index, item in enumerate(self.import_field_dict.items()):
            items = list(item)
            key = items[0]
            value = items[1]
            if isinstance(value, dict):
                header_data.append(value.get("title"))
                hidden_header.append(value.get('display'))
                choices = value.get("choices", {})
                if choices.get("data"):
                    data_list = []
                    data_list.extend(choices.get("data").keys())
                    validation_data_dict[value.get("title")] = data_list
                elif choices.get("queryset") and choices.get("values_name"):
                    data_list = choices.get("queryset").values_list(choices.get("values_name"), flat=True)
                    validation_data_dict[value.get("title")] = list(data_list)
                else:
                    continue
                column_letter = get_column_letter(len(validation_data_dict))
                dv = DataValidation(
                    type="list",
                    formula1=f"{quote_sheetname('data')}!${column_letter}$2:${column_letter}${len(validation_data_dict[value.get('title')]) + 1}",
                    allow_blank=True,
                )
                ws.add_data_validation(dv)
                dv.add(f"{get_column_letter(index + 3)}2:{get_column_letter(index + 3)}1048576")
            else:
                header_data.append(value)
                hidden_header.append(key)
        # Add a column of data
        ws1.append(list(validation_data_dict.keys()))
        for index, validation_data in enumerate(validation_data_dict.values()):
            for inx, ele in enumerate(validation_data):
                ws1[f"{get_column_letter(index + 1)}{inx + 2}"] = ele
        #--------
        df_len_max = [self.get_string_len(ele) for ele in header_data]
        row = get_column_letter(len(hidden_header) + 1)
        column = 1
        ws.append(header_data)
        for index, results in enumerate(data):
            results_list = []
            for h_index, h_item in enumerate(hidden_header):
                for key, val in results.items():
                    if key == h_item:
                        if val is None or val == "":
                            results_list.append("")
                        elif isinstance(val,list):
                            results_list.append(str(val))
                        else:
                            results_list.append(val)
                        # Calculate the maximum column width
                        if isinstance(val,str):
                            result_column_width = self.get_string_len(val)
                            if h_index != 0 and result_column_width > df_len_max[h_index]:
                                df_len_max[h_index] = result_column_width
            ws.append([index+1,*results_list])
            column += 1
        # 　Update column width
        for index, width in enumerate(df_len_max):
            ws.column_dimensions[get_column_letter(index + 1)].width = width
        tab = Table(displayName="Table", ref=f"A1:{row}{column}")  # Name Manager
        style = TableStyleInfo(
            name="TableStyleLight11",
            showFirstColumn=True,
            showLastColumn=True,
            showRowStripes=True,
            showColumnStripes=True,
        )
        tab.tableStyleInfo = style
        ws.add_table(tab)
        wb.save(response)
        return response


class ExportSerializerMixin:
    """
    Custom export function
    """

    # Export fields
    export_field_label = []
    # Export serializer
    export_serializer_class = None
    # Maximum width of table header，default50Characters
    export_column_width = 50

    def is_number(self,num):
        try:
            float(num)
            return True
        except ValueError:
            pass

        try:
            import unicodedata
            unicodedata.numeric(num)
            return True
        except (TypeError, ValueError):
            pass
        return False

    def get_string_len(self, string):
        """
        Get the maximum length of the string
        :param string:
        :return:
        """
        length = 4
        if string is None:
            return length
        if self.is_number(string):
            return length
        for char in string:
            length += 2.1 if ord(char) > 256 else 1
        return round(length, 1) if length <= self.export_column_width else self.export_column_width

    @action(methods=['get'],detail=False)
    def export_data(self, request: Request, *args, **kwargs):
        """
        Export function
        :param request:
        :param args:
        :param kwargs:
        :return:
        """
        queryset = self.filter_queryset(self.get_queryset())
        assert self.export_field_label, "'%s' Please configure the corresponding export template field。" % self.__class__.__name__
        assert self.export_serializer_class, "'%s' Please configure the corresponding export serializer。" % self.__class__.__name__
        data = self.export_serializer_class(queryset, many=True, request=request).data
        try:
            async_export_data.delay(
                data,
                str(f"Export{get_verbose_name(queryset)}-{datetime.datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"),
                DownloadCenter.objects.create(creator=request.user, task_name=f'{get_verbose_name(queryset)}Data Export Task', dept_belong_id=request.user.dept_id).pk,
                self.export_field_label
            )
            return SuccessResponse(msg="Import task has been created，Please go‘Download Center’Waiting for download")
        except:
            pass
        # Exportexcel surface
        response = HttpResponse(content_type="application/msexcel")
        response["Access-Control-Expose-Headers"] = f"Content-Disposition"
        response["content-disposition"] = f'attachment;filename={quote(str(f"Export{get_verbose_name(queryset)}.xlsx"))}'
        wb = Workbook()
        ws = wb.active
        header_data = ["Serial number", *self.export_field_label.values()]
        hidden_header = ["#", *self.export_field_label.keys()]
        df_len_max = [self.get_string_len(ele) for ele in header_data]
        row = get_column_letter(len(self.export_field_label) + 1)
        column = 1
        ws.append(header_data)
        for index, results in enumerate(data):
            results_list = []
            for h_index, h_item in enumerate(hidden_header):
                for key,val in results.items():
                    if key == h_item:
                        if val is None or val=="":
                            results_list.append("")
                        else:
                            results_list.append(val)
                        # Calculate the maximum column width
                        result_column_width = self.get_string_len(val)
                        if h_index !=0 and result_column_width > df_len_max[h_index]:
                            df_len_max[h_index] = result_column_width
            ws.append([index + 1, *results_list])
            column += 1
        # 　Update column width
        for index, width in enumerate(df_len_max):
            ws.column_dimensions[get_column_letter(index + 1)].width = width
        tab = Table(displayName="Table", ref=f"A1:{row}{column}")  # Name Manager
        style = TableStyleInfo(
            name="TableStyleLight11",
            showFirstColumn=True,
            showLastColumn=True,
            showRowStripes=True,
            showColumnStripes=True,
        )
        tab.tableStyleInfo = style
        ws.add_table(tab)
        wb.save(response)
        return response
